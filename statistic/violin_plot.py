import openpyxl
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np


# Load the Excel file
workbook = openpyxl.load_workbook('statistic-survey987744.xlsx')

# Select the sheet you want to scrape data from
worksheet = workbook['results-survey987744']

# Initialize variables to hold scores for each method
E_scores = [[], [], [], [], []]
PR_scores = [[], [], [], [], []]
RP_scores = [[], [], [], [], []]
critera_dict = {"Melody": 0, "Rythm": 1, "Artistic Intention": 2,  "Repetition" : 3, "Subjective rating" : 4}

def violin_plot(E_scores, PR_scores, RP_scores):

    methods = ["Event Based", "Relative Pitch", "Piano Roll"]
    criteria = ["Melody", "Rhythm", "Artistic Intention", "Repetition", "Subjective rating"]

    fig, axs = plt.subplots(nrows=1, ncols=len(methods), figsize=(50, 8))

    # Set the color palette
    palette = "husl"

    # Plot each violin plot
    for i, met in enumerate(methods):
        if met == "Event Based":
            data = E_scores
        elif met == "Relative Pitch":
            data = RP_scores
        else:
            data = PR_scores

        # Plot the violin plot with updated styling
        sns.violinplot(data=data, ax=axs[i], palette=palette, edgecolor='white', linewidth=2)



        # Set the title for the subplot
        axs[i].set_title(f"{met}", fontsize=16,)

        # Set the tick label font size
        axs[i].tick_params(axis='both', which='major', labelsize=16)

        # Remove the top and right spines
        axs[i].spines['top'].set_visible(False)
        axs[i].spines['right'].set_visible(False)

        # Set the x and y labels
        axs[i].set_xlabel("")
        axs[i].set_ylabel("Score", fontsize=14)

        # Set the grid
        axs[i].grid(axis='y')


        # Set the x tick labels
        axs[i].set_xticklabels(criteria, rotation=45)
        
        majors = [1, 2, 3, 4, 5, 6, 7]
        axs[i].yaxis.set_major_locator(ticker.FixedLocator(majors))


    # Adjust the spacing
    plt.subplots_adjust(wspace=0.5)
    plt.show()

def find_scores(criteria, method_name, idx):
    first_row = idx+4
    column = 2
    criteria = critera_dict[criteria]
    for i in range(1,8):
        row = first_row + i
        cell = worksheet.cell(row, column)
        score = cell.value
        for j in range(score):
            if method_name == 'E':
                E_scores[criteria].append(i)
            elif method_name == 'PR':
                PR_scores[criteria].append(i)
            elif method_name == 'RP':
                RP_scores[criteria].append(i)
    
    
i = 0
for idx, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=7, values_only=True)):
    print(row[0])
    if row[0] and 'C' not in str(row[0]):
        if 'E' in str(row[0]) or 'PR' in str(row[0]) or 'RP' in str(row[0]):
            i += 1
            method_name = row[0]
            if 'E' in row[0]:
                method_name = 'E'
            if 'PR' in row[0]:
                method_name = 'PR'
            if 'RP' in row[0]:
                method_name = 'RP'

            criteria = row[0][row[0].find('[')+1:row[0].find(']')]
            find_scores(criteria, method_name, idx)
            
        

            
            

    
                
                


                
print('E scores:', E_scores[4]+E_scores[0]+E_scores[2]+E_scores[1])
print('PR scores:', PR_scores[4]+PR_scores[0]+PR_scores[2]+PR_scores[1])
print('RP scores:', RP_scores[4]+RP_scores[0]+RP_scores[2]+RP_scores[1])
violin_plot(E_scores, PR_scores, RP_scores)


